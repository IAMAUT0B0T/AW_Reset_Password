from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import time
import random
import pyotp
import string

s = Service('D:\BotWax\chromedriver.exe')
driver = None
actions = None
wait = None
num_rows = 0    
def initialize_driver():
    options = webdriver.ChromeOptions()
    chrome_options = Options()
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-plugins")
    chrome_options.add_argument("--disable-popup-blocking")
    chrome_options.add_argument("--disable-background-networking")
    chrome_options.add_argument("--disable-sync")
    chrome_options.add_argument("--disable-translate")
    chrome_options.add_argument("--disable-desktop-notifications")
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    global driver
    driver = webdriver.Chrome(service=s, options=options)
    driver.set_window_size(1024, 768)
    global actions
    actions = ActionChains(driver)
    global wait
    wait = WebDriverWait(driver, 30)

def close_driver():
    global driver
    if driver:
        driver.quit()

def get_accounts():
    workbook = load_workbook('WAX_accounts.xlsx')
    worksheet = workbook.active
    accounts = worksheet.iter_rows(values_only=True)
    next(accounts)  # Skip the first row (header)
    workbook.close()
    global num_rows 
    num_rows = worksheet.max_row-1
    return accounts
    
def login(email, password, twofa_secret):
    
    driver.get("https://www.mycloudwallet.com/signin")
    time.sleep(2)
    try:
    # Wait for the element to be visible
        input_email = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="root"]/div[3]/div[2]/div[1]/div[1]/div[2]/div[5]/input')))
        input_email.clear()
        input_email.send_keys(email)
    except Exception as e:
        print("Element not found within specified time:", e)

    try:
        # Wait for the element to be visible
        input_password = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="root"]/div[3]/div[2]/div[1]/div[1]/div[2]/div[7]/span/input')))
        input_password.clear()
        input_password.send_keys(password)
    except Exception as e:
        print("Element not found within specified time:", e)
    time.sleep(1)
    try:
        element = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[3]/div[2]/div[1]/div[1]/div[2]/div[9]/div[2]/div/button')))
        actions.double_click(element).perform()
    except Exception as e:
        print("Element is not clickable:", e)
    
    #2fa
    try:
        input_2facode = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="tfacode"]')))
        input_2facode.clear()
        otp = pyotp.TOTP(twofa_secret)
        input_2facode.send_keys(otp.now())
    except Exception as e:
        print("Element not found within specified time:", e)

    time.sleep(1)
    try:
        element = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[2]/div[2]/div[1]/div/div/div[4]/div/button')))
        element.click()
    except Exception as e:
        print("Element is not clickable:", e)
    time.sleep(2)

def logout():
    driver.get("https://www.mycloudwallet.com/dashboard")
    time.sleep(2)

    try:
        element = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[2]/div[3]/div/div/div[1]/img')))
        element.click()
    except Exception as e:
        print("Element is not clickable:", e)
    time.sleep(2)
    try:
        element = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div/div/div[2]/div/div/div[4]/img')))
        element.click()
    except Exception as e:
        print("Element is not clickable:", e)
    time.sleep(3)


def generate_password(length=16):
    # Define character sets
    uppercase_chars = string.ascii_uppercase
    lowercase_chars = string.ascii_lowercase
    digits = string.digits
    special_chars = ".,@#$%&+-"

    # Ensure at least one character from each category
    password = [random.choice(uppercase_chars),
                random.choice(lowercase_chars),
                random.choice(digits),
                random.choice(special_chars)]

    # Fill the remaining length - 4 with random characters
    password.extend(random.choice(string.ascii_letters + digits + special_chars) for _ in range(length - 4))

    # Shuffle the password
    random.shuffle(password)

    # Ensure at least one special character is included
    password.append(random.choice(special_chars))

    # Shuffle the password again to ensure randomness
    random.shuffle(password)

    return ''.join(password)

def reset_password(old_password, new_password):
    driver.get("https://www.mycloudwallet.com/settings/basic-information")
    time.sleep(2)
    try:
        element = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/div[3]/div/div[3]/div[2]/div/div[2]/div/div/div[2]/div/div[3]/div/div/div[3]/div/button')))
        element.click()
    except Exception as e:
        print("Element is not clickable:", e)
    
    try:
        # Wait for the element to be visible
        input_password = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="current-password"]')))
        input_password.clear()
        input_password.send_keys(old_password)
    except Exception as e:
        print("Element not found within specified time:", e)

    try:
        # Wait for the element to be visible
        input_password = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="new-password"]')))
        input_password.clear()
        input_password.send_keys(new_password)
    except Exception as e:
        print("Element not found within specified time:", e)

    try:
        # Wait for the element to be visible
        input_password = wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="confirm-password"]')))
        input_password.clear()
        input_password.send_keys(new_password)
    except Exception as e:
        print("Element not found within specified time:", e)

    time.sleep(1)
    
    try:
        element = wait.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div/div[2]/div/div[2]/div/div[2]/form/div[2]/div/div/div/div/div/div[2]/div/button')))
        actions.double_click(element).perform()
    except Exception as e:
        print("Element is not clickable:", e)

    status_updatepassword = ""
    try:
        element = wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[2]/div/div/div/div/div/div[2]/div')))
        status_updatepassword = element.text
    except Exception as e:
        print("Element is not clickable:", e)

    if(status_updatepassword == "Your password has been updated successfully."):
        print("Password reset successfully!!")
    else:
        print("Failed to reset password.")
        exit()
    time.sleep(2)

def update_user(username, new_password):
    workbook = load_workbook('WAX_accounts.xlsx')
    worksheet = workbook.active
    found = False
    # Assuming the usernames are in column B (2nd column)
    for cell in worksheet['B']:
        if cell.value == username:
            row_number = cell.row
            worksheet.cell(row=row_number, column=4, value=new_password)  # Assuming password is in column 4
            found = True
            break

    if found:
        print("Password updated successfully for user:", username)
    else:
        print("User not found:", username)
    workbook.save("WAX_accounts.xlsx")
    workbook.close()

def logging(username, old_password, new_password):
    # Open the file in write mode, which creates the file if it doesn't exist
    with open('log.txt', 'a') as file:
    # Write some text to the file
        file.write("Username: "+username+" Old Password: "+old_password+" New Password: "+new_password+"\n")
def main():
    logo = """
    ___        __ ____                _       ____                                     _ 
   / \ \      / /|  _ \ ___  ___  ___| |_    |  _ \ __ _ ___ _____      _____  _ __ __| |
  / _ \ \ /\ / / | |_) / _ \/ __|/ _ \ __|   | |_) / _` / __/ __\ \ /\ / / _ \| '__/ _` |
 / ___ \ V  V /  |  _ <  __/\__ \  __/ |_    |  __/ (_| \__ \__ \\ V  V / (_) | | | (_| |
/_/   \_\_/\_/___|_| \_\___||___/\___|\__|___|_|   \__,_|___/___/ \_/\_/ \___/|_|  \__,_|
            |_____|                     |_____|                                          
 
 AW_Reset_Password_v1 by AUTOBOT.
 https://www.facebook.com/people/Autobot/61558488084551/                                                       
            """
    print(logo)
    accounts = get_accounts()
    
    for account in accounts:
        initialize_driver()
        username = account[1]
        print("["+str(account[0])+"/"+str(num_rows)+"]", "Username:", username)
        email = account[2]
        old_password = account[3]
        print("Old Password:", old_password)
        twofa_secret = account[4]
        new_password = generate_password()
        print("New Password:", new_password)
        login(email, old_password, twofa_secret)
        reset_password(old_password, new_password)
        logging(username, old_password, new_password)
        update_user(username, new_password)
        close_driver()
        time.sleep(2)
    exit()
if __name__ == "__main__":
    main()
